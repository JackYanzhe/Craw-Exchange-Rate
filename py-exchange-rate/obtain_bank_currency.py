import datetime
import json
import random
import time
import decimal
import openpyxl
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/67.0.3396.79 Safari/537.36'
}
# 中银汇率地址
url = 'https://srh.bankofchina.com/search/whpj/search_cn.jsp'


def obtainCurrencyRateHtml(param,url):
    try:
        html = requests.post(url=url, data=param, headers=headers)
        html.encoding = html.apparent_encoding
        return html.text
        # print(html.text)
    except Exception as e:
        print('接口调用异常：', e)
    return ''


# 传入时间获取相应汇率信息
def obtainCurrencyRate(date):
    data = []
    # '','','','','','','','',''
    param_arr = [['英镑','GBP'],['港币','HKD'],['美元','USD'],['新加坡元','SGD'],['日元','JPY'],['加拿大元','CAD'],['澳大利亚元','AUD'],['欧元','EUR'],['印尼卢比','IDR'],['印度卢比','INR']]
    try:
        for currency in param_arr:
            params = {
                'erectDate': date,
                'nothing': date,
                'pjname': currency[0]
            }
            currency_map = {'中行汇率时间':date,'币种':currency[1]}
            htmlText = obtainCurrencyRateHtml(params,url)
            # print(htmlText)
            soup = BeautifulSoup(htmlText, 'html.parser')
            currencyTable = soup.find(class_='BOC_main publish').find(name='table')
            # print(currencyTable.text)
            currency_trs = currencyTable.find_all(name='tr')
            result = json.loads(json.dumps(data))
            cur_titles = currency_trs[0].find_all(name='th')
            cur_values = currency_trs[1].find_all(name='td')
            for i in range(len(cur_titles)):
                cur_title = cur_titles[i].text
                cur_value = cur_values[i].text
                currency_map[cur_title] = cur_value
            print(currency,currency_map)
            data.append(currency_map)
            randomTime = random.randint(500,2300)/1000
            print('等待时间：',str(randomTime))
            time.sleep(randomTime)
    except Exception as e:
        print('接口调用异常：', e)
    return data


# 读
def readExcel(path,sheet,rowSheet):
    data_list = []
    wb = load_workbook(filename=path)
    sheet = wb[sheet]
    sheetList = sheet[rowSheet]
    for item in sheetList:
        data_list.append(item.value)
    # 实例调用save方法可以关闭表格，否则后续调用会出现占用报错Permission denied
    wb.save(path)
    return data_list


# 写/修改
def writeExcel(data, path, sheetname):
    # 实例化excel表
    wb = openpyxl.Workbook()
    wb.create_sheet(sheetname)
    # 修改_active_sheet_index值可以更换要激活的sheet页索引，默认是0
    wb._active_sheet_index = 1
    sheet = wb.active
    # 循环插入sheet页
    sheet.cell(row=1, column=1, value=str('中行汇率时间'))
    sheet.cell(row=1, column=2, value=str('货币名称'))
    sheet.cell(row=1, column=3, value=str('币种'))
    sheet.cell(row=1, column=4, value=str('中行折算价(RMB)'))
    sheet.cell(row=1, column=5, value=str('中行折算价'))
    sheet.cell(row=1, column=6, value=str('最终发布时间'))
    for num in range(len(data)):
        sheet.cell(row=num+2, column=1, value=str(data[num].get('中行汇率时间')))
        sheet.cell(row=num+2, column=2, value=str(data[num].get('货币名称')))
        sheet.cell(row=num+2, column=3, value=str(data[num].get('币种')))
        priceStr = data[num].get('中行折算价')
        price = divBigdecimal(100,float(priceStr),6)
        sheet.cell(row=num+2, column=4, value=str(price))
        sheet.cell(row=num+2, column=5, value=str(data[num].get('中行折算价')))
        sheet.cell(row=num+2, column=6, value=str(data[num].get('发布时间')))
    # 指定保存路径
    wb.save(path)
    return "保存成功"


# 两数相除时精确保留位数，避免丢失精度
def divBigdecimal(num1,num2,byteNum):
    # 保留位数
    decimal.getcontext().prec = byteNum
    division = decimal.Decimal(num1) / decimal.Decimal(num2)
    return division


if __name__ == '__main__':
    # detester = '2020-07-27'
    dateList = readExcel("E:\\Data\\tem-data-py.xlsx",'Sheet1','A')
    resultData = []
    for date in dateList:
        restDt = obtainCurrencyRate(date)
        for dt in restDt:
            resultData.append(dt)
        print('当前时间完成汇率拉取：',resultData)
    rt = writeExcel(resultData, "E:\\Data\\rate-data.xlsx", "Sheet")
    print(rt)
